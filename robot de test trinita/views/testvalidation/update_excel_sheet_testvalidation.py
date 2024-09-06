from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def update_excel_sheet_testvalidation(v1: float, v2: float, testvalidation, fichier_excel: str = 'dosexcel/adherent.xlsx'):
    # Charger le classeur Excel
    wb = load_workbook(fichier_excel)
    
    # Sélectionner la feuille 'testvalidation'
    sheet = wb['testvalidation']
    
    # Trouver la colonne 'numéro contrat'
    numero_contrat_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'numero_contrat':
            numero_contrat_col = col
            break
    
    if numero_contrat_col is None:
        raise ValueError("La colonne 'numéro contrat' n'a pas été trouvée.")
    
    # Trouver la ligne correspondant au numéro de contrat
    target_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=numero_contrat_col).value == testvalidation.numero_contrat:
            target_row = row
            break
    
    if target_row is None:
        raise ValueError(f"Le numéro de contrat '{testvalidation.numero_contrat}' n'a pas été trouvé.")
    
    # Trouver ou créer les colonnes 'v1' et 'v2'
    v1_col = None
    v2_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'resutlat_trouveé_en_€':
            v1_col = col
        elif sheet.cell(row=1, column=col).value == 'resultat_calculeé_TTC_en_€':
            v2_col = col
    
    # Créer les colonnes 'v1' et 'v2' si elles n'existent pas
    if v1_col is None:
        v1_col = sheet.max_column + 1
        sheet.cell(row=1, column=v1_col, value='resutlat_trouveé_en_€')
    
    if v2_col is None:
        v2_col = sheet.max_column + 1 if v1_col != sheet.max_column + 1 else sheet.max_column + 2
        sheet.cell(row=1, column=v2_col, value='resultat_calculeé_TTC_en_€')
    
    # Définir les couleurs pour le remplissage
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Vert
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')    # Rouge
    
    # Écrire la valeur v1 et colorier la cellule
    v1_cell = sheet.cell(row=target_row, column=v1_col, value=v1)
    if v1 > testvalidation.solde_attendu:
        v1_cell.fill = green_fill
    else:
        v1_cell.fill = red_fill
    
    # Écrire la valeur v2 et colorier la cellule
    v2_cell = sheet.cell(row=target_row, column=v2_col, value=v2)
    if v2 > testvalidation.solde_attendu:
        v2_cell.fill = green_fill
    else:
        v2_cell.fill = red_fill
    
    # Sauvegarder les modifications
    wb.save(fichier_excel)
    print(f"Les valeurs ont été mises à jour pour le contrat {testvalidation.numero_contrat} avec la coloration appropriée.")
