import pandas as pd
from openpyxl import load_workbook, Workbook  # Corrected import
import os
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def ajouter_num_adhe_num_contrat_resultatfeuilleexcel(excel_path, sheet_name, numero_adherent, statut_adherent, numero_contrat, statut_contrat, res, commentaire):
    # Définir les couleurs
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Vert
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rouge
    grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gris

    try:
        # Charger le fichier Excel
        if not os.path.exists(excel_path):  # Check if file exists
            raise FileNotFoundError

        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            # Charger la feuille existante
            sheet = book[sheet_name]
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
        else:
            # Créer une nouvelle feuille avec des colonnes
            df = pd.DataFrame(columns=['numero_adherent', 'statut_adherent', 'numero_contrat', 'statut_contrat', 'resultat', 'commentaire'])
            sheet = book.create_sheet(sheet_name)
    except FileNotFoundError:
        # Créer un nouveau fichier Excel avec la feuille si le fichier n'existe pas
        book = Workbook()
        sheet = book.active
        sheet.title = sheet_name
        df = pd.DataFrame(columns=['numero_adherent', 'statut_adherent', 'numero_contrat', 'statut_contrat', 'resultat', 'commentaire'])
    
    # Ajouter une nouvelle ligne avec le numéro d'adhérent
    new_row = pd.DataFrame([{'numero_adherent': numero_adherent, 'statut_adherent':statut_adherent , 'numero_contrat': numero_contrat, 'statut_contrat':statut_contrat, 'resultat': res, 'commentaire':commentaire}])
    df = pd.concat([df, new_row], ignore_index=True)

    # Sauvegarder le DataFrame dans le fichier Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1:  # Skip the header row for coloring
                fill_color = green_fill if df.at[r_idx-2, 'resultat'] == 'OK' else red_fill
                cell.fill = fill_color
            else:
                cell.fill = grey_fill
    book.save(excel_path)
